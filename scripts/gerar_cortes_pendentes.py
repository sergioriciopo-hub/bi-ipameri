import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

pend   = pd.read_parquet("C:/BI_Ipameri/data/pendencia_atual.parquet")
bairro = pd.read_parquet("C:/BI_Ipameri/data/dim_bairro.parquet")
categ  = pd.read_parquet("C:/BI_Ipameri/data/dim_categoria.parquet")

cortes = pend[pend["fl_corte_pendente"] == True].copy()
cortes = cortes.merge(bairro[["id_bairro","nm_bairro_dim"]], on="id_bairro", how="left")
cortes = cortes.merge(categ[["id_categoria","nm_rsf_categoria_dim"]], on="id_categoria", how="left")
cortes["dt_vencimento"] = pd.to_datetime(cortes["dt_vencimento"])
cortes = cortes.sort_values("vl_divida", ascending=False).reset_index(drop=True)

por_faixa  = cortes.groupby("faixa_atraso").agg(Qtd=("id_localizacao","count"), Divida=("vl_divida","sum")).reset_index().sort_values("faixa_atraso")
por_bairro = cortes.groupby("nm_bairro_dim").agg(Qtd=("id_localizacao","count"), Divida=("vl_divida","sum")).reset_index().sort_values("Qtd", ascending=False)

wb = Workbook()

# Estilos
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HDR_FILL   = PatternFill("solid", start_color="1A6FAD")
TOTAL_FONT = Font(name="Arial", bold=True, size=11)
TOTAL_FILL = PatternFill("solid", start_color="D6E4F0")
DATA_FONT  = Font(name="Arial", size=10)
GRAY_FILL  = PatternFill("solid", start_color="F2F2F2")
CENTER     = Alignment(horizontal="center", vertical="center")
LEFT       = Alignment(horizontal="left",   vertical="center")
RIGHT      = Alignment(horizontal="right",  vertical="center")
thin       = Side(style="thin", color="CCCCCC")
BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)
BRL        = '#,##0.00'
DATE_FMT   = 'DD/MM/YYYY'

def hdr_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = CENTER; c.border = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width

def apply(c, font=None, fill=None, align=None, fmt=None, border=None):
    if font:   c.font = font
    if fill:   c.fill = fill
    if align:  c.alignment = align
    if fmt:    c.number_format = fmt
    if border: c.border = border

# ABA 1 — Cortes Pendentes
ws = wb.active
ws.title = "Cortes Pendentes"
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 22

headers = [
    ("Localizacao", 14), ("Bairro", 34), ("Categoria", 18),
    ("Dias Atraso", 13), ("Faixa de Atraso", 22), ("Divida (R$)", 16), ("Vencimento", 14),
]
for i, (label, width) in enumerate(headers, 1):
    hdr_cell(ws, 1, i, label, width)

for idx, row in cortes.iterrows():
    excel_row = idx + 2
    fill = GRAY_FILL if idx % 2 == 0 else None
    values = [
        (row["id_localizacao"],       None,     CENTER),
        (row["nm_bairro_dim"],        None,     LEFT),
        (row["nm_rsf_categoria_dim"], None,     LEFT),
        (row["nr_dias_atraso"],       None,     CENTER),
        (row["faixa_atraso"],         None,     LEFT),
        (row["vl_divida"],            BRL,      RIGHT),
        (row["dt_vencimento"],        DATE_FMT, CENTER),
    ]
    for col, (val, fmt, align) in enumerate(values, 1):
        c = ws.cell(row=excel_row, column=col, value=val)
        c.font = DATA_FONT; c.border = BORDER; c.alignment = align
        if fmt:  c.number_format = fmt
        if fill: c.fill = fill

n = len(cortes)
total_row = n + 2

# Linha de totais
for col in range(1, 8):
    c = ws.cell(row=total_row, column=col)
    c.font = TOTAL_FONT; c.fill = TOTAL_FILL; c.border = BORDER; c.alignment = CENTER

ws.cell(row=total_row, column=1).value = "TOTAL"
ws.cell(row=total_row, column=2).value = str(n) + " registros"
ws.cell(row=total_row, column=2).alignment = LEFT
c_tot = ws.cell(row=total_row, column=6)
c_tot.value = "=SUM(F2:F{})".format(n + 1)
c_tot.number_format = BRL
c_tot.alignment = RIGHT

# ABA 2 — Resumo
ws2 = wb.create_sheet("Resumo")
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 18

t = ws2.cell(row=1, column=1, value="Cortes Pendentes - Aguas de Ipameri")
t.font = Font(name="Arial", bold=True, size=14, color="1A6FAD")
ws2.merge_cells("A1:C1")

ws2.cell(row=2, column=1, value="Gerado em: {}".format(date.today().strftime("%d/%m/%Y")))
ws2.cell(row=2, column=1).font = Font(name="Arial", size=9, color="888888")
ws2.merge_cells("A2:C2")

kpi_data = [
    ("Total de Cortes Pendentes", n,                        None),
    ("Divida Total (R$)",         cortes["vl_divida"].sum(), BRL),
    ("Divida Media (R$)",         cortes["vl_divida"].mean(), BRL),
]
for i, (label, val, fmt) in enumerate(kpi_data, 4):
    cl = ws2.cell(row=i, column=1, value=label)
    cl.font = Font(name="Arial", bold=True, size=10)
    cl.fill = PatternFill("solid", start_color="EBF5FB")
    cl.border = BORDER; cl.alignment = LEFT
    cv = ws2.cell(row=i, column=2, value=val)
    cv.font = Font(name="Arial", bold=True, size=11, color="1A6FAD")
    cv.border = BORDER; cv.alignment = RIGHT
    if fmt: cv.number_format = fmt

# Por faixa
r = 9
for col, lbl in [(1,"Faixa de Atraso"), (2,"Qtd"), (3,"Divida (R$)")]:
    c = ws2.cell(row=r, column=col, value=lbl)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER; c.alignment = CENTER

for i, row in por_faixa.reset_index(drop=True).iterrows():
    rr = r + 1 + i
    fill = GRAY_FILL if i % 2 == 0 else None
    for col, val, fmt, align in [
        (1, row["faixa_atraso"], None, LEFT),
        (2, row["Qtd"],          None, CENTER),
        (3, row["Divida"],       BRL,  RIGHT),
    ]:
        c = ws2.cell(row=rr, column=col, value=val)
        c.font = DATA_FONT; c.border = BORDER; c.alignment = align
        if fmt:  c.number_format = fmt
        if fill: c.fill = fill

# Por bairro
r2 = r + len(por_faixa) + 3
for col, lbl in [(1,"Bairro"), (2,"Qtd"), (3,"Divida (R$)")]:
    c = ws2.cell(row=r2, column=col, value=lbl)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER; c.alignment = CENTER

for i, row in por_bairro.reset_index(drop=True).iterrows():
    rr = r2 + 1 + i
    fill = GRAY_FILL if i % 2 == 0 else None
    for col, val, fmt, align in [
        (1, row["nm_bairro_dim"], None, LEFT),
        (2, row["Qtd"],           None, CENTER),
        (3, row["Divida"],        BRL,  RIGHT),
    ]:
        c = ws2.cell(row=rr, column=col, value=val)
        c.font = DATA_FONT; c.border = BORDER; c.alignment = align
        if fmt:  c.number_format = fmt
        if fill: c.fill = fill

out = "C:/BI_Ipameri/Apresentacoes/Cortes_Pendentes_2026-05-11.xlsx"
wb.save(out)
print("Salvo:", out)
print("{} registros | R$ {:,.2f}".format(len(cortes), cortes["vl_divida"].sum()))
