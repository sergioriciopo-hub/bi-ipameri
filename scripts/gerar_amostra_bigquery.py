import os
from google.cloud import bigquery
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

KEY_PATH = r"C:\Users\SérgioRiciopo\AGUAS DE IPAMERI\03 COMERCIAL - Documentos\01 - Projetos e Propostas\Jtech\DATA SET\ipameri type service account_28.04.2026.json"
PROJECT = "br-ist-jtech-clientes"
DATASET = "jtechbi_ipameri"
OUTPUT = r"C:\BI_Ipameri\amostra_bigquery_abril2026.xlsx"

os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = KEY_PATH
client = bigquery.Client(project=PROJECT)

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Arial", size=9)
ALT_FILL = PatternFill("solid", start_color="F2F2F2")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DATE_TYPES = {"TIMESTAMP", "DATE", "DATETIME"}


def detect_date_col(schema):
    for f in schema:
        if f.field_type in DATE_TYPES:
            return f.name
    return None


def fetch_sample(table_ref, schema):
    date_col = detect_date_col(schema)
    full = f"`{PROJECT}.{DATASET}.{table_ref}`"
    if date_col:
        q = f"""
        SELECT * FROM {full}
        WHERE CAST({date_col} AS DATE) BETWEEN '2026-04-01' AND '2026-04-30'
        LIMIT 50
        """
        df = client.query(q).to_dataframe()
        if len(df) > 0:
            return df, date_col, "Abril/2026"
        q2 = f"SELECT * FROM {full} ORDER BY {date_col} DESC LIMIT 20"
        df = client.query(q2).to_dataframe()
        return df, date_col, "Mais recentes"
    else:
        q = f"SELECT * FROM {full} LIMIT 30"
        df = client.query(q).to_dataframe()
        return df, None, "Sem filtro de data"


def style_sheet(ws, df):
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row, 1):
            if hasattr(val, 'isoformat'):
                val = val.isoformat()
            elif pd.isna(val) if not isinstance(val, str) else False:
                val = None
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = BORDER
            if fill:
                cell.fill = fill

    for col_idx, col_name in enumerate(df.columns, 1):
        lengths = [len(str(col_name))] + [len(str(v)) for v in df.iloc[:, col_idx-1].astype(str)]
        max_len = max(lengths) if lengths else 10
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def safe_sheet_name(name):
    name = name[:31]
    for ch in r'\/*?:[]\x00':
        name = name.replace(ch, '_')
    return name


print("Listando tabelas do dataset...")
tables = list(client.list_tables(f"{PROJECT}.{DATASET}"))
print(f"Total encontrado: {len(tables)} tabelas")

wb = Workbook()
idx_ws = wb.active
idx_ws.title = "ÍNDICE"

idx_headers = ["#", "Tabela", "Total Colunas", "Colunas de Data", "Linhas Amostra", "Filtro Aplicado", "Observação"]
for col_idx, h in enumerate(idx_headers, 1):
    cell = idx_ws.cell(row=1, column=col_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
idx_ws.freeze_panes = "A2"

idx_row = 2
for i, table in enumerate(tables, 1):
    tname = table.table_id
    print(f"[{i}/{len(tables)}] {tname}...", end=" ")
    try:
        tref = client.get_table(f"{PROJECT}.{DATASET}.{tname}")
        schema = tref.schema
        date_cols = [f.name for f in schema if f.field_type in DATE_TYPES]
        df, date_col, filtro = fetch_sample(tname, schema)
        print(f"{len(df)} linhas ({filtro})")

        sheet_name = safe_sheet_name(tname)
        ws = wb.create_sheet(title=sheet_name)
        if len(df) > 0:
            style_sheet(ws, df)
        else:
            ws["A1"] = "Sem dados disponíveis"
            ws["A1"].font = Font(name="Arial", italic=True, color="888888")

        obs = ""
        idx_ws.cell(row=idx_row, column=1, value=i)
        idx_ws.cell(row=idx_row, column=2, value=tname)
        idx_ws.cell(row=idx_row, column=3, value=len(schema))
        idx_ws.cell(row=idx_row, column=4, value=", ".join(date_cols) if date_cols else "—")
        idx_ws.cell(row=idx_row, column=5, value=len(df))
        idx_ws.cell(row=idx_row, column=6, value=filtro)
        idx_ws.cell(row=idx_row, column=7, value=obs)
    except Exception as e:
        print(f"ERRO: {e}")
        idx_ws.cell(row=idx_row, column=1, value=i)
        idx_ws.cell(row=idx_row, column=2, value=tname)
        idx_ws.cell(row=idx_row, column=7, value=f"Erro: {str(e)[:80]}")
    for col in range(1, 8):
        idx_ws.cell(row=idx_row, column=col).font = Font(name="Arial", size=9)
        idx_ws.cell(row=idx_row, column=col).border = BORDER
        if idx_row % 2 == 0:
            idx_ws.cell(row=idx_row, column=col).fill = ALT_FILL
    idx_row += 1

idx_ws.column_dimensions["A"].width = 5
idx_ws.column_dimensions["B"].width = 35
idx_ws.column_dimensions["C"].width = 15
idx_ws.column_dimensions["D"].width = 30
idx_ws.column_dimensions["E"].width = 15
idx_ws.column_dimensions["F"].width = 18
idx_ws.column_dimensions["G"].width = 50
idx_ws.row_dimensions[1].height = 30

wb.save(OUTPUT)
print(f"\nPlanilha salva em:\n{OUTPUT}")
