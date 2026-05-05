"""
ETL para extrair dados de Desconto de Energia
Lê arquivos Excel de "Controles de consumo de Energia Eletrica"
e exporta em formato Parquet com análise de descontos
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime

SHAREPOINT_PATH = Path(r"C:\Users\SérgioRiciopo\AGUAS DE IPAMERI\04 ADM-FINANCEIRO - Documentos\Energia Elétrica\Controles de consumo de Energia Eletrica")
OUTPUT_PATH = Path(__file__).parent.parent / "data"

MESES = {
    "Janeiro": 1, "Fevereiro": 2, "Março": 3, "Abril": 4,
    "Maio": 5, "Junho": 6, "Julho": 7, "Agosto": 8,
    "Setembro": 9, "Outubro": 10, "Novembro": 11, "Dezembro": 12
}

def extrair_desconto_arquivo(filepath, ano, fornecedor):
    """Extrai dados de desconto de um arquivo Excel"""
    print(f"Processando: {filepath.name} ({fornecedor})")

    wb = openpyxl.load_workbook(filepath)

    if fornecedor not in wb.sheetnames:
        print(f"  Aba '{fornecedor}' não encontrada")
        return []

    ws = wb[fornecedor]
    dados = []

    # Ler linha 2 para encontrar os meses (coluna com mês)
    meses_cols = {}  # {mes_num: col_idx}
    for col in range(1, 50):
        cell_val = ws.cell(row=2, column=col).value
        if cell_val and cell_val in MESES:
            meses_cols[MESES[cell_val]] = col

    if not meses_cols:
        print(f"  Nenhum mês encontrado na aba {fornecedor}")
        return []

    # Ler linhas (começando da linha 4, que tem os dados)
    for row_idx in range(4, 100):
        uc_cell = ws.cell(row=row_idx, column=1).value
        if not uc_cell or str(uc_cell).startswith("TOTAL"):
            break

        uc = str(uc_cell).strip()

        # Processar cada mês encontrado
        for mes_num, col_base in meses_cols.items():
            # Colunas: col_base = Valor Pago, col_base+1 = Valor Cheio, col_base+2 = Desconto
            valor_pago = ws.cell(row=row_idx, column=col_base).value
            valor_cheio = ws.cell(row=row_idx, column=col_base + 1).value
            desconto = ws.cell(row=row_idx, column=col_base + 2).value

            # Verificar se tem dados
            if valor_pago is None:
                continue

            # Converter valores
            try:
                valor_pago = float(valor_pago) if isinstance(valor_pago, (int, float)) else 0
                valor_cheio = float(valor_cheio) if isinstance(valor_cheio, (int, float)) else valor_pago
                desconto = float(desconto) if isinstance(desconto, (int, float)) else 0

                if valor_pago == 0:
                    continue

                # Calcular % de desconto
                pct_desconto = (desconto / valor_cheio * 100) if valor_cheio > 0 else 0

                dados.append({
                    "uc": uc,
                    "mes": mes_num,
                    "ano": ano,
                    "mes_ano": f"{mes_num:02d}/{ano}",
                    "valor_pago": valor_pago,
                    "valor_cheio": valor_cheio,
                    "desconto_r": desconto,
                    "desconto_pct": pct_desconto,
                    "fornecedor": fornecedor,
                    "data_extracao": datetime.now().isoformat()
                })
            except (ValueError, TypeError):
                continue

    return dados

def main():
    """Processa todos os arquivos e exporta Parquet"""
    print("Iniciando ETL de Desconto de Energia...\n")

    todas_dados = []

    # Arquivos e anos
    arquivos = {
        "CONSUMO ENERGIA ELÉTRICA 2024.xlsx": 2024,
        "CONSUMO ENERGIA ELÉTRICA 2025.xlsx": 2025,
        "CONSUMO ENERGIA ELÉTRICA 2026.xlsx": 2026,
    }

    for arquivo, ano in arquivos.items():
        filepath = SHAREPOINT_PATH / arquivo
        if not filepath.exists():
            print(f"Arquivo não encontrado: {arquivo}\n")
            continue

        for fornecedor in ["Matrix", "EchoEnergia"]:
            dados = extrair_desconto_arquivo(filepath, ano, fornecedor)
            todas_dados.extend(dados)
            print(f"  Extraído {len(dados)} registros de {fornecedor}\n")

    if todas_dados:
        df = pd.DataFrame(todas_dados)

        # Estatísticas
        print("\n" + "="*60)
        print("RESUMO DA EXTRAÇÃO")
        print("="*60)
        print(f"Total de registros: {len(df)}")
        print(f"\nPor fornecedor:")
        print(df.groupby("fornecedor").size())
        print(f"\nPor ano:")
        print(df.groupby("ano").size())
        print(f"\nDescontos médios:")
        print(df.groupby("fornecedor")["desconto_pct"].agg(["mean", "min", "max"]))

        # Exportar Parquet
        output_file = OUTPUT_PATH / "desconto_energia.parquet"
        df.to_parquet(output_file, index=False)
        print(f"\n[OK] Exportado: {output_file}")
    else:
        print("Nenhum dado extraído!")

if __name__ == "__main__":
    main()
